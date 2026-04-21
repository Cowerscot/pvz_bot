# avito_core.py

import time
from pathlib import Path
from datetime import datetime, timedelta
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
from openpyxl import load_workbook
from collections import defaultdict
from calendar import monthrange

from full_config import REPORTS_DIR, AVITO_URL


def wait_for_authorization(driver, logger, timeout=300):
    """Ожидание авторизации Авито с автонажатием кнопки входа"""
    logger.info("  🔐 Проверка авторизации...")

    # Сразу пробуем нажать кнопку «Войти»
    try:
        submit_btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, "submit-button"))
        )
        url_before = driver.current_url
        submit_btn.click()
        logger.info("  🔘 Кнопка входа нажата, ожидаем переход...")
        time.sleep(3)

        # Если URL изменился — авторизация прошла
        if driver.current_url != url_before and "/gw/login" not in driver.current_url:
            logger.info("  ✅ Авторизован (автовход)")
            return True
    except Exception:
        pass  # Кнопки нет — страница уже авторизована или форма другая

    # Проверяем — вдруг уже авторизованы
    if "/gw/login" not in driver.current_url and "pvz.avito.ru" in driver.current_url:
        logger.info("  ✅ Авторизован")
        return True

    # Ждём ручной авторизации
    logger.info("  ⏳ Ожидание ручной авторизации...")
    start = time.time()
    while time.time() - start < timeout:
        if "/gw/login" not in driver.current_url and "pvz.avito.ru" in driver.current_url:
            logger.info("  ✅ Авторизован")
            return True
        time.sleep(2)
    
    logger.error("  ❌ Таймаут авторизации")
    return False


def download_avito_report(driver, logger, avito_dir):
    """Скачивание отчёта Авито
    
    Args:
        driver: WebDriver
        logger: Логгер
        avito_dir: Папка для сохранения (Авито)
    
    Returns:
        Path: Путь к скачанному файлу или None
    """
    logger.info("  📥 Скачивание отчёта...")

    result_file = None
    try:
        # Прямой переход в раздел аналитики
        analytics_url = "https://pvz.avito.ru/analytics"
        if driver.current_url.rstrip('/') != analytics_url.rstrip('/'):
            logger.info(f"  🌐 Переход: {analytics_url}")
            driver.get(analytics_url)
            time.sleep(3)
            logger.info(f"  ✅ Текущий URL: {driver.current_url}")
        
        # Кнопка скачивания
        download_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((
                By.CSS_SELECTOR,
                "[class*='DownloadTableButton']"
            ))
        )
        
        # Удаляем только analytics.xlsx (дефолтное имя от Авито) перед скачиванием
        for known_name in ["analytics.xlsx", "analytics (1).xlsx", "analytics (2).xlsx"]:
            stale = avito_dir / known_name
            if stale.exists():
                try:
                    stale.unlink()
                    logger.debug(f"  🗑️ Удалён старый analytics файл: {stale.name}")
                except Exception as e:
                    logger.debug(f"  ⚠️ Не удалось удалить {stale.name}: {e}")

        # Запоминаем файлы до скачивания
        initial_files = set(avito_dir.glob("*.xlsx"))
        
        download_btn.click()
        time.sleep(2)
        
        # Ожидание появления нового файла
        for i in range(30):
            time.sleep(1)

            # Проверка временных файлов xlsx (игнорируем .crdownload от htm-мусора)
            temp_files = [f for f in avito_dir.glob("*.crdownload") if "htm" not in f.name.lower()]
            temp_files += [f for f in avito_dir.glob("*.tmp") if "htm" not in f.name.lower()]
            if temp_files:
                logger.info(f"  ⏳ Временный xlsx файл: {[f.name for f in temp_files]}")
                continue

            current_files = set(avito_dir.glob("*.xlsx"))
            new_files = current_files - initial_files
            
            if new_files:
                new_file = max(new_files, key=lambda f: f.stat().st_mtime)
                if new_file.stat().st_size > 0:
                    logger.info(f"  ✅ Файл скачан: {new_file.name}")
                    result_file = new_file
                    break
                continue

            # analytics.xlsx мог быть перезаписан — проверяем по времени
            analytics_file = avito_dir / "analytics.xlsx"
            if analytics_file.exists() and analytics_file.stat().st_size > 0:
                if analytics_file.stat().st_mtime >= (time.time() - i - 5):
                    logger.info(f"  ✅ Найден analytics.xlsx (перезаписан)")
                    result_file = analytics_file
                    break

        if result_file is None:
            # Последний шанс — берём analytics.xlsx если есть
            analytics_file = avito_dir / "analytics.xlsx"
            if analytics_file.exists() and analytics_file.stat().st_size > 0:
                logger.info(f"  ✅ Найден analytics.xlsx (fallback)")
                result_file = analytics_file
            else:
                all_files = list(avito_dir.iterdir())
                logger.error(f"  ❌ Таймаут скачивания файла. Файлы в папке: {[f.name for f in all_files]}")

        return result_file

    except Exception as e:
        logger.error(f"  ❌ Ошибка скачивания: {e}")
        return None

    finally:
        # Удаляем мусорные htm/crdownload файлы которые браузер скачивает вместе с отчётом
        for pattern in ["*.htm", "*.html", "*.crdownload", "*.tmp"]:
            for junk in avito_dir.glob(pattern):
                try:
                    junk.unlink(missing_ok=True)
                    logger.debug(f"  🗑️ Удалён мусор: {junk.name}")
                except Exception as e:
                    logger.debug(f"  ⚠️ Не удалось удалить {junk.name}: {e}")


def get_avito_monthly_totals(filepath, logger):
    """Получение сумм и количества дней по ПВЗ
    
    Args:
        filepath: Путь к файлу Excel
        logger: Логгер
    
    Returns:
        tuple: ({адрес: сумма_за_месяц}, {адрес: количество_дней})
    """
    logger.info("  📊 Анализ месячных данных...")
    
    try:
        wb = load_workbook(filepath, data_only=True)
        
        target_sheet = None
        for sheet in wb.worksheets:
            if sheet.title == "Расширенная":
                target_sheet = sheet
                break
        
        if not target_sheet:
            logger.warning("  ⚠️ Лист 'Расширенная' не найден")
            return {}, {}
        
        yesterday = (datetime.now() - timedelta(days=1)).date()
        month_start = yesterday.replace(day=1)
        
        target_sheet.delete_rows(1, 3)
        
        pvz_totals = defaultdict(float)
        pvz_dates = defaultdict(set)  # Уникальные даты для каждого ПВЗ
        
        for row in target_sheet.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 10:
                continue
            
            pvz_name = row[1]
            date_val = row[2]
            amount_val = row[9]
            
            if not pvz_name or not amount_val:
                continue
            
            try:
                if isinstance(date_val, datetime):
                    row_date = date_val.date()
                elif isinstance(date_val, str):
                    for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
                        try:
                            row_date = datetime.strptime(date_val, fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        continue
                else:
                    continue
                
                if month_start <= row_date <= yesterday:
                    amount = float(amount_val)
                    pvz_name_str = str(pvz_name)
                    pvz_totals[pvz_name_str] += amount
                    pvz_dates[pvz_name_str].add(row_date)
                
            except (ValueError, TypeError):
                continue
        
        totals = {pvz: round(amount) for pvz, amount in pvz_totals.items()}
        days_count = {pvz: len(dates) for pvz, dates in pvz_dates.items()}
        first_day = {pvz: min(dates).day for pvz, dates in pvz_dates.items()}
        
        logger.info(f"  📊 Данных за месяц: {len(totals)} ПВЗ")
        for pvz in totals:
            logger.info(f"    {pvz}: {days_count[pvz]} дней, первый день: {first_day[pvz]}")

        return totals, days_count, first_day
        
    except Exception as e:
        logger.error(f"  ❌ Ошибка анализа месячных данных: {e}")
        return {}, {}, {}


def analyze_avito_report(filepath, logger):
    """Анализ Excel отчёта Авито за вчерашний день
    
    Args:
        filepath: Путь к файлу Excel
        logger: Логгер
    
    Returns:
        dict: Данные по ПВЗ {адрес: сумма}
    """
    logger.info("  📊 Анализ отчёта...")
    
    try:
        wb = load_workbook(filepath, data_only=True)
        
        # Ищем лист "Расширенная" (без "таблица")
        target_sheet = None
        for sheet in wb.worksheets:
            if sheet.title == "Расширенная":
                target_sheet = sheet
                break
        
        if not target_sheet:
            logger.warning("  ⚠️ Лист 'Расширенная' не найден")
            logger.info(f"  Доступные листы: {[s.title for s in wb.worksheets]}")
            return {}
        
        # Вчерашняя дата для фильтрации
        yesterday = (datetime.now() - timedelta(days=1)).date()
        logger.info(f"  📅 Ищем данные за: {yesterday.strftime('%d.%m.%Y')}")
        
        # Удаляем первые 3 строки
        target_sheet.delete_rows(1, 3)
        
        # Колонки: B=ПВЗ (индекс 1), C=Дата (индекс 2), J=Сумма (индекс 9)
        pvz_totals = defaultdict(float)
        rows_checked = 0
        matched_rows = 0
        
        for row in target_sheet.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 10:
                continue
            
            rows_checked += 1
            
            pvz_name = row[1]     # Колонка B (Название ПВЗ)
            date_val = row[2]     # Колонка C (Дата)
            amount_val = row[9]   # Колонка J (Сумма)
            
            if not pvz_name or not amount_val:
                continue
            
            # Фильтр по вчерашней дате
            try:
                if isinstance(date_val, datetime):
                    row_date = date_val.date()
                elif isinstance(date_val, str):
                    # Пробуем разные форматы
                    for fmt in ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y']:
                        try:
                            row_date = datetime.strptime(date_val, fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        continue
                else:
                    continue
                
                if row_date != yesterday:
                    continue
                
                matched_rows += 1
                
                amount = float(amount_val)
                pvz_totals[str(pvz_name)] += amount
                
            except (ValueError, TypeError) as e:
                logger.debug(f"  Ошибка парсинга строки: {e}")
                continue
        
        logger.info(f"  📋 Всего строк: {rows_checked}, совпадений: {matched_rows}")
        
        if not pvz_totals:
            logger.warning(f"  ⚠️ Нет данных за {yesterday.strftime('%d.%m.%Y')}")
            return {}
        
        # Округляем суммы
        result = {pvz: round(amount) for pvz, amount in pvz_totals.items()}
        
        # Логируем итоги
        total = sum(result.values())
        logger.info(f"  💰 Всего ПВЗ: {len(result)} | Сумма за день: {total:,.0f} ₽")
        for pvz, amount in sorted(result.items()):
            logger.info(f"    {pvz}: {amount:,.0f} ₽")
        
        return result
        
    except Exception as e:
        logger.error(f"  ❌ Ошибка анализа: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise


def process_avito_report(driver, logger):
    """Основная функция обработки отчёта Авито
    
    Returns:
        dict: {
            'daily': {адрес: сумма за вчера},
            'forecast': {адрес: прогноз на месяц},
            'last_date': date
        }
    """
    avito_dir = REPORTS_DIR / "Авито"
    avito_dir.mkdir(parents=True, exist_ok=True)
    
    # Проверяем существующий файл ЗА СЕГОДНЯ
    today = datetime.now().date()
    final_path = avito_dir / f"{today.day:02d}.{today.month:02d}.{today.year}.xlsx"
    
    if final_path.exists():
        logger.info(f"  ✅ Файл существует: {final_path.name}")
        # Анализируем существующий файл
        yesterday = (datetime.now() - timedelta(days=1)).date()
        pvz_data = analyze_avito_report(final_path, logger)
        
        # Получаем суммы, количество дней и первый рабочий день за месяц
        monthly_totals, days_count, first_day = get_avito_monthly_totals(final_path, logger)

        # Среднее за день = сумма_за_месяц / количество_дней_ПВЗ
        avg_data = {pvz: round(total / days_count[pvz]) if days_count.get(pvz, 0) > 0 else 0
                    for pvz, total in monthly_totals.items()}

        # Прогноз = среднее × рабочие дни ПВЗ в месяце
        days_in_month = monthrange(today.year, today.month)[1]
        forecast = {pvz: round(avg * (days_in_month - (first_day.get(pvz, 1) - 1)))
                    for pvz, avg in avg_data.items()}

        if forecast:
            total_forecast = sum(forecast.values())
            logger.info(f"  📈 Прогноз на месяц: {total_forecast:,.0f} ₽")
        
        logger.info("✅ Авито")
        
        return {
            'daily': pvz_data,
            'avg': avg_data,
            'forecast': forecast,
            'last_date': yesterday
        }
    
    # Файла нет - скачиваем
    driver.execute_cdp_cmd("Page.setDownloadBehavior", {
        "behavior": "allow",
        "downloadPath": str(avito_dir)
    })
    
    try:
        logger.info(f"  🌐 Переход: {AVITO_URL}")
        driver.get(AVITO_URL)
        logger.info(f"  ✅ URL загружен, текущий: {driver.current_url}")
        time.sleep(3)
        
        if "/gw/login" in driver.current_url:
            if not wait_for_authorization(driver, logger):
                raise Exception("Ошибка авторизации Авито")
        
        # Скачивание отчёта
        downloaded_file = download_avito_report(driver, logger, avito_dir)
        if not downloaded_file:
            raise Exception("Не удалось скачать отчёт")
        
        # Переименовываем в дату вчера
        if downloaded_file.exists():
            downloaded_file.rename(final_path)
        
        # Удаляем лишние файлы
        for pattern in ["downloads.*", "*.htm", "*.html"]:
            for junk in avito_dir.glob(pattern):
                try:
                    junk.unlink(missing_ok=True)
                    logger.debug(f"  🗑️ Удален: {junk.name}")
                except Exception as e:
                    logger.debug(f"  ⚠️ Не удалось удалить {junk.name}: {e}")
        
        # Анализ за вчерашний день
        yesterday = (datetime.now() - timedelta(days=1)).date()
        pvz_data = analyze_avito_report(final_path, logger)
        
        # Получаем суммы, количество дней и первый рабочий день за месяц
        monthly_totals, days_count, first_day = get_avito_monthly_totals(final_path, logger)

        # Среднее за день = сумма_за_месяц / количество_дней_ПВЗ
        avg_data = {pvz: round(total / days_count[pvz]) if days_count.get(pvz, 0) > 0 else 0
                    for pvz, total in monthly_totals.items()}

        # Прогноз = среднее × рабочие дни ПВЗ в месяце
        days_in_month = monthrange(today.year, today.month)[1]
        forecast = {pvz: round(avg * (days_in_month - (first_day.get(pvz, 1) - 1)))
                    for pvz, avg in avg_data.items()}

        if forecast:
            total_forecast = sum(forecast.values())
            logger.info(f"  📈 Прогноз на месяц: {total_forecast:,.0f} ₽")
        
        logger.info("✅ Авито")
        
        return {
            'daily': pvz_data,
            'avg': avg_data,
            'forecast': forecast,
            'last_date': yesterday
        }
        
    except Exception as e:
        logger.error(f"Ошибка обработки Авито: {e}")
        raise
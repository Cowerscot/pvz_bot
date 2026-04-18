# yandex_core.py

import time
from pathlib import Path
from datetime import datetime, timedelta
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
from openpyxl import load_workbook
from collections import defaultdict
from calendar import monthrange

from full_config import YANDEX_AUTH_URL, YANDEX_REPORT_URL, REPORTS_DIR


def ensure_authorized(driver, logger):
    """Проверка авторизации Яндекс"""
    logger.info("  🔐 Проверка авторизации...")
    original_handles = set(driver.window_handles)
    driver.execute_script("window.open('');")
    new_handle = [h for h in driver.window_handles if h not in original_handles][0]
    driver.switch_to.window(new_handle)
    driver.get(YANDEX_AUTH_URL)
    time.sleep(2)
    
    current_url = driver.current_url
    
    if "passport.yandex.ru" not in current_url and "auth" not in current_url.lower():
        logger.info("  ✅ Авторизован")
    else:
        try:
            account_btn = driver.find_element(By.CSS_SELECTOR, ".UserLogin-displayName")
            account_btn.click()
            time.sleep(2)
        except:
            pass
        
        try:
            buttons = driver.find_elements(By.XPATH, "//button[contains(., '@')] | //a[contains(., '@')]")
            if buttons:
                buttons[0].click()
                time.sleep(2)
        except:
            pass
        
        if "passport.yandex.ru" in driver.current_url or "auth" in driver.current_url.lower():
            logger.info("  ⏳ Ожидание авторизации...")
            for _ in range(90):
                time.sleep(2)
                if "passport.yandex.ru" not in driver.current_url and "auth" not in driver.current_url.lower():
                    logger.info("  ✅ Авторизован")
                    break
    
    driver.close()
    driver.switch_to.window(list(driver.window_handles)[0])


def open_report_page_and_download(driver, logger, download_dir):
    """Открытие страницы отчётов и скачивание"""
    logger.info("  📄 Открытие страницы отчётов...")
    
    # Устанавливаем папку загрузки через CDP
    driver.execute_cdp_cmd('Page.setDownloadBehavior', {
        'behavior': 'allow',
        'downloadPath': str(download_dir)
    })
    
    driver.execute_script("window.open('');")
    driver.switch_to.window(driver.window_handles[-1])
    driver.get(YANDEX_REPORT_URL)
    time.sleep(3)
    
    if "passport" in driver.current_url.lower() or "auth" in driver.current_url.lower():
        time.sleep(15)
        driver.get(YANDEX_REPORT_URL)
        time.sleep(2)
    
    wait = WebDriverWait(driver, 30)
    
    download_btn = None
    strategies = [
        lambda: wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Скачать')]"))),
        lambda: driver.find_element(By.XPATH, "//button[contains(., 'Скачать')]"),
    ]
    
    for i, strategy in enumerate(strategies, 1):
        try:
            download_btn = strategy()
            break
        except:
            continue
    
    if not download_btn:
        raise Exception("Кнопка 'Скачать' не найдена")
    
    if not download_btn.is_displayed():
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", download_btn)
        time.sleep(1)
    
    try:
        download_btn.click()
    except:
        driver.execute_script("arguments[0].click();", download_btn)
    
    logger.info("  ✅ Скачивание запущено")


def wait_for_xlsx_download(download_dir, logger, timeout=30):
    """Ожидание скачивания XLSX файла"""
    logger.info("  ⏳ Ожидание файла...")
    main_dir = download_dir
    
    initial = {f for f in main_dir.glob("*.xlsx") if f.is_file()}
    waited = 0
    
    while waited < timeout:
        time.sleep(2)
        waited += 2
        current = {f for f in main_dir.glob("*.xlsx") if f.is_file()}
        new_files = current - initial
        
        if new_files:
            new_file = max(new_files, key=lambda f: f.stat().st_ctime)
            time.sleep(2)
            logger.info(f"  ✅ Получен: {new_file.name}")
            return new_file
    
    raise Exception(f"Файл не появился за {timeout} секунд")


def analyze_report(filepath, logger):
    """Анализ Excel отчёта Яндекс с группировкой по ID ПВЗ"""
    logger.info(f"  📊 Анализ отчёта...")
    wb = load_workbook(filepath, read_only=True)
    if "Транзакции" not in wb.sheetnames:
        raise ValueError("Лист 'Транзакции' не найден")
    
    sheet = wb["Транзакции"]
    headers = [cell.value for cell in sheet[1]]
    
    try:
        pvz_id_col = 0  # Первая колонка - ID ПВЗ
        time_col = headers.index("Время (мск)")
        amount_col = headers.index("Стоимость услуги, руб")
    except ValueError:
        raise ValueError("Не найдены колонки: 'Время (мск)' или 'Стоимость услуги, руб'")
    
    # Группировка: {pvz_id: {date: amount}}
    pvz_daily_totals = defaultdict(lambda: defaultdict(float))
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        pvz_id = row[pvz_id_col]
        time_val = row[time_col]
        amount_val = row[amount_col]
        
        if not pvz_id or not time_val or not amount_val:
            continue
        
        # Парсим дату
        if isinstance(time_val, datetime):
            day = time_val.date()
        elif isinstance(time_val, str):
            try:
                dt = datetime.strptime(time_val, "%Y-%m-%d %H:%M:%S")
                day = dt.date()
            except ValueError:
                continue
        else:
            continue
        
        # Парсим сумму
        if isinstance(amount_val, (int, float)):
            amount = float(amount_val)
        elif isinstance(amount_val, str):
            try:
                amount = float(amount_val.replace(',', '.'))
            except ValueError:
                continue
        else:
            continue
        
        pvz_daily_totals[pvz_id][day] += amount
    
    if not pvz_daily_totals:
        raise ValueError("Нет данных для анализа")
    
    # Рассчитываем метрики для каждого ПВЗ
    result = {
        'pvz_data': {},
        'last_date': None
    }
    
    overall_last_date = None
    
    for pvz_id, daily_totals in pvz_daily_totals.items():
        last_date = max(daily_totals.keys())
        last_amount = round(daily_totals[last_date])
        
        # Обновляем общую последнюю дату
        if overall_last_date is None or last_date > overall_last_date:
            overall_last_date = last_date
        
                # Среднее считаем только по доступным дням текущего месяца
        month_days = {k: v for k, v in daily_totals.items()
                      if k.month == last_date.month and k.year == last_date.year}

        available_days = len(month_days)
        avg_daily = round(sum(month_days.values()) / available_days) if available_days > 0 else 0

        # Прогноз = среднее × рабочие дни ПВЗ в месяце
        # (от первого рабочего дня ПВЗ до конца месяца)
        days_in_month = monthrange(last_date.year, last_date.month)[1]
        first_working_day = min(month_days.keys()).day if month_days else 1
        working_days_in_month = days_in_month - (first_working_day - 1)
        forecast = round(avg_daily * working_days_in_month)
        
        result['pvz_data'][pvz_id] = {
            'last_date': last_date,
            'last_amount': last_amount,
            'avg_daily': avg_daily,
            'forecast': forecast
        }
    
    result['last_date'] = overall_last_date
    
    # Логируем итоги по каждому ПВЗ
    for pvz_id, data in sorted(result['pvz_data'].items()):
        logger.info(f"  💰 ID_{pvz_id}: {data['last_amount']:,} ₽ | Прогноз: {data['forecast']:,} ₽")
    
    return result


def process_yandex_report(driver, logger):
    """Основная функция обработки отчёта Яндекс"""
    yandex_dir = REPORTS_DIR / "Яндекс"
    yandex_dir.mkdir(parents=True, exist_ok=True)
    
    today = datetime.now().date()
    final_path = yandex_dir / f"{today.day:02d}.{today.month:02d}.{today.year}.xlsx"
    
    # Проверяем существующий файл за сегодня
    if final_path.exists():
        logger.info(f"  ✅ Файл существует: {final_path.name}")
        report_data = analyze_report(final_path, logger)
        return report_data
    
    # Файла нет - скачиваем
    try:
        # Авторизация
        ensure_authorized(driver, logger)
        time.sleep(2)
        
                        # Скачивание отчёта
        open_report_page_and_download(driver, logger, yandex_dir)
        
        # Ожидание файла в папке Яндекс
        new_file = wait_for_xlsx_download(yandex_dir, logger)
        
        if new_file.exists():
            new_file.rename(final_path)
        
        # Анализ
        report_data = analyze_report(final_path, logger)
        
        # Закрываем вкладку Яндекса
        if len(driver.window_handles) > 1:
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
        
        return report_data
        
    except Exception as e:
        logger.error(f"Ошибка обработки Яндекс: {e}")
        # Закрываем вкладку при ошибке
        if len(driver.window_handles) > 1:
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
        raise